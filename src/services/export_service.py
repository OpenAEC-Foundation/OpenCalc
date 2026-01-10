"""
ExportService - Export naar Excel en ODS formaten
"""

from pathlib import Path
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from odf.opendocument import OpenDocumentSpreadsheet, OpenDocumentText
    from odf.table import Table, TableRow, TableCell, TableColumn
    from odf.text import P, H, Span
    from odf.style import Style, TextProperties, TableCellProperties, TableColumnProperties, ParagraphProperties
    HAS_ODF = True
except ImportError:
    HAS_ODF = False


class ExportService:
    """Service voor het exporteren van begrotingen"""

    def __init__(self, schedule):
        self._schedule = schedule

    def export_xlsx(self, file_path: str) -> bool:
        """Exporteer naar Excel XLSX formaat"""
        if not HAS_OPENPYXL:
            print("openpyxl niet geinstalleerd. Installeer met: pip install openpyxl")
            return False

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Begroting"

            # Styles
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
            header_font_white = Font(bold=True, size=12, color="FFFFFF")
            chapter_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            chapter_font = Font(bold=True)
            currency_format = '€ #,##0.00'
            thin_border = Border(
                left=Side(style='thin', color='E2E8F0'),
                right=Side(style='thin', color='E2E8F0'),
                top=Side(style='thin', color='E2E8F0'),
                bottom=Side(style='thin', color='E2E8F0')
            )

            # Titel
            ws.merge_cells('A1:F1')
            ws['A1'] = self._schedule.name
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal='center')

            # Headers
            headers = ['Code', 'Omschrijving', 'Eenheid', 'Hoeveelheid', 'Prijs', 'Totaal']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # Data
            row = 4
            for item in self._schedule.items:
                # Hoofdstuk
                ws.cell(row=row, column=1, value=item.identification).font = chapter_font
                ws.cell(row=row, column=2, value=item.name).font = chapter_font
                ws.cell(row=row, column=6, value=item.subtotal).number_format = currency_format
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = chapter_fill
                    ws.cell(row=row, column=col).border = thin_border
                row += 1

                # Kostenposten
                for child in item.children:
                    ws.cell(row=row, column=1, value=child.identification).border = thin_border
                    ws.cell(row=row, column=2, value=child.name).border = thin_border
                    ws.cell(row=row, column=3, value=child.unit_symbol).border = thin_border
                    ws.cell(row=row, column=4, value=child.quantity).border = thin_border
                    ws.cell(row=row, column=5, value=child.unit_price).number_format = currency_format
                    ws.cell(row=row, column=5).border = thin_border
                    ws.cell(row=row, column=6, value=child.subtotal).number_format = currency_format
                    ws.cell(row=row, column=6).border = thin_border
                    row += 1

            # Totaal
            row += 1
            ws.cell(row=row, column=5, value="Subtotaal:").font = Font(bold=True)
            ws.cell(row=row, column=6, value=self._schedule.subtotal).number_format = currency_format
            ws.cell(row=row, column=6).font = Font(bold=True)

            row += 1
            ws.cell(row=row, column=5, value=f"BTW ({self._schedule.vat_rate}%):").font = Font(bold=True)
            ws.cell(row=row, column=6, value=self._schedule.vat_amount).number_format = currency_format

            row += 1
            ws.cell(row=row, column=5, value="Totaal:").font = Font(bold=True, size=12)
            ws.cell(row=row, column=6, value=self._schedule.total).number_format = currency_format
            ws.cell(row=row, column=6).font = Font(bold=True, size=12)

            # Kolom breedtes
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 45
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 14

            wb.save(file_path)
            return True

        except Exception as e:
            print(f"Excel export error: {e}")
            return False

    def export_ods(self, file_path: str) -> bool:
        """Exporteer naar ODS formaat (LibreOffice)"""
        if not HAS_ODF:
            print("odfpy niet geinstalleerd. Installeer met: pip install odfpy")
            return False

        try:
            doc = OpenDocumentSpreadsheet()

            # Styles
            bold_style = Style(name="Bold", family="table-cell")
            bold_style.addElement(TextProperties(fontweight="bold"))
            doc.styles.addElement(bold_style)

            # Table
            table = Table(name="Begroting")

            # Titel
            row = TableRow()
            cell = TableCell()
            cell.addElement(P(text=self._schedule.name))
            row.addElement(cell)
            table.addElement(row)

            # Lege rij
            table.addElement(TableRow())

            # Headers
            row = TableRow()
            for header in ['Code', 'Omschrijving', 'Eenheid', 'Hoeveelheid', 'Prijs', 'Totaal']:
                cell = TableCell()
                cell.addElement(P(text=header))
                row.addElement(cell)
            table.addElement(row)

            # Data
            for item in self._schedule.items:
                # Hoofdstuk
                row = TableRow()
                row.addElement(TableCell(valuetype="string"))
                cell = TableCell()
                cell.addElement(P(text=item.identification))
                row.addElement(cell)
                cell = TableCell()
                cell.addElement(P(text=item.name))
                row.addElement(cell)
                for _ in range(3):
                    row.addElement(TableCell())
                cell = TableCell(valuetype="currency", currency="EUR", value=str(item.subtotal))
                row.addElement(cell)
                table.addElement(row)

                # Kostenposten
                for child in item.children:
                    row = TableRow()
                    cell = TableCell()
                    cell.addElement(P(text=child.identification))
                    row.addElement(cell)
                    cell = TableCell()
                    cell.addElement(P(text=child.name))
                    row.addElement(cell)
                    cell = TableCell()
                    cell.addElement(P(text=child.unit_symbol or ""))
                    row.addElement(cell)
                    cell = TableCell(valuetype="float", value=str(child.quantity))
                    row.addElement(cell)
                    cell = TableCell(valuetype="currency", currency="EUR", value=str(child.unit_price))
                    row.addElement(cell)
                    cell = TableCell(valuetype="currency", currency="EUR", value=str(child.subtotal))
                    row.addElement(cell)
                    table.addElement(row)

            doc.spreadsheet.addElement(table)
            doc.save(file_path)
            return True

        except Exception as e:
            print(f"ODS export error: {e}")
            return False

    def export_odt(self, file_path: str) -> bool:
        """Exporteer naar ODT formaat (LibreOffice Writer)"""
        if not HAS_ODF:
            print("odfpy niet geinstalleerd. Installeer met: pip install odfpy")
            return False

        try:
            doc = OpenDocumentText()

            # Titel style
            title_style = Style(name="Title", family="paragraph")
            title_style.addElement(TextProperties(fontsize="18pt", fontweight="bold"))
            title_style.addElement(ParagraphProperties(marginbottom="0.5cm"))
            doc.styles.addElement(title_style)

            # Chapter style
            chapter_style = Style(name="Chapter", family="paragraph")
            chapter_style.addElement(TextProperties(fontsize="12pt", fontweight="bold"))
            chapter_style.addElement(ParagraphProperties(margintop="0.3cm", marginbottom="0.1cm"))
            doc.styles.addElement(chapter_style)

            # Normal style
            normal_style = Style(name="Normal", family="paragraph")
            normal_style.addElement(TextProperties(fontsize="10pt"))
            doc.styles.addElement(normal_style)

            # Bold style
            bold_style = Style(name="Bold", family="paragraph")
            bold_style.addElement(TextProperties(fontsize="10pt", fontweight="bold"))
            doc.styles.addElement(bold_style)

            # Titel
            title = H(outlinelevel=1, stylename=title_style, text=self._schedule.name)
            doc.text.addElement(title)

            # Subtitel met datum
            from datetime import datetime
            subtitle = P(stylename=normal_style, text=f"Datum: {datetime.now().strftime('%d-%m-%Y')}")
            doc.text.addElement(subtitle)

            # Lege regel
            doc.text.addElement(P(text=""))

            # Tabel met kolom definities
            table = Table(name="Begroting")

            # Headers
            row = TableRow()
            for header in ['Code', 'Omschrijving', 'Eenheid', 'Hoeveelheid', 'Prijs', 'Totaal']:
                cell = TableCell()
                p = P(stylename=bold_style, text=header)
                cell.addElement(p)
                row.addElement(cell)
            table.addElement(row)

            # Data
            for item in self._schedule.items:
                # Hoofdstuk
                row = TableRow()
                cell = TableCell()
                cell.addElement(P(stylename=chapter_style, text=item.identification or ""))
                row.addElement(cell)
                cell = TableCell()
                cell.addElement(P(stylename=chapter_style, text=item.name))
                row.addElement(cell)
                for _ in range(3):
                    row.addElement(TableCell())
                cell = TableCell()
                cell.addElement(P(stylename=chapter_style, text=f"€ {item.subtotal:,.2f}".replace(",", ".")))
                row.addElement(cell)
                table.addElement(row)

                # Kostenposten
                for child in item.children:
                    row = TableRow()
                    cell = TableCell()
                    cell.addElement(P(stylename=normal_style, text=child.identification or ""))
                    row.addElement(cell)
                    cell = TableCell()
                    cell.addElement(P(stylename=normal_style, text=child.name))
                    row.addElement(cell)
                    cell = TableCell()
                    cell.addElement(P(stylename=normal_style, text=child.unit_symbol or ""))
                    row.addElement(cell)
                    cell = TableCell()
                    cell.addElement(P(stylename=normal_style, text=f"{child.quantity:,.2f}".replace(",", ".")))
                    row.addElement(cell)
                    cell = TableCell()
                    cell.addElement(P(stylename=normal_style, text=f"€ {child.unit_price:,.2f}".replace(",", ".")))
                    row.addElement(cell)
                    cell = TableCell()
                    cell.addElement(P(stylename=normal_style, text=f"€ {child.subtotal:,.2f}".replace(",", ".")))
                    row.addElement(cell)
                    table.addElement(row)

            doc.text.addElement(table)

            # Totalen
            doc.text.addElement(P(text=""))
            doc.text.addElement(P(stylename=bold_style, text=f"Subtotaal: € {self._schedule.subtotal:,.2f}".replace(",", ".")))
            doc.text.addElement(P(stylename=normal_style, text=f"BTW ({self._schedule.vat_rate}%): € {self._schedule.vat_amount:,.2f}".replace(",", ".")))
            doc.text.addElement(P(stylename=bold_style, text=f"Totaal: € {self._schedule.total:,.2f}".replace(",", ".")))

            doc.save(file_path)
            return True

        except Exception as e:
            print(f"ODT export error: {e}")
            return False
